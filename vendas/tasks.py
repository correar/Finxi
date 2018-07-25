# Create your tasks here
from __future__ import absolute_import, unicode_literals
from celery import shared_task
from .models import Empresa, Produto, Categoria
from openpyxl import load_workbook

@shared_task
def read_upload_file(f):
    wb = load_workbook(filename=f, read_only=True)
    sheetName = wb.sheetnames[0]
    ws = wb[sheetName]
    for row in ws.iter_rows(row_offset=1):
        produto_nome = row[0].value
        categoria_nome = row[1].value
        unidades = row[2].value
        pc = row[3].value
        tv = row[4].value
        pc = pc.split(" ")
        preco_custo = pc[1].replace(",",".")
        tv = tv.split(" ")
        total_venda = tv[1].replace(",",".")
        c = Categoria.objects.create(categoria_nome=categoria_nome)
        c.save()
        p = Produto.objects.create(produto_nome=produto_nome,preco_custo=preco_custo,unidades=unidades,total_venda=total_venda,categorias=c)
        p.save()
        e = Empresa.objects.last()
        p.empresas.add(e.pk)
